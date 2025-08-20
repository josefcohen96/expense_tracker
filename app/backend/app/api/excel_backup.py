from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any
import logging

from openpyxl import Workbook, load_workbook

LOG = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).resolve().parents[2]  # .../expense_tracker/app
EXCEL_DIR = ROOT_DIR / "backups" / "excel"
EXCEL_DIR.mkdir(parents=True, exist_ok=True)

# Adjust header fields to match your expense dict keys
HEADERS = ["id", "date", "amount", "category", "notes", "created_at", "updated_at"]


def _parse_date(value: Optional[str]) -> datetime:
    if not value:
        return datetime.utcnow()
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(value)
    except Exception:
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(value, fmt)
            except Exception:
                continue
    return datetime.utcnow()


def _file_for_dt(dt: datetime) -> Path:
    return EXCEL_DIR / f"expenses_{dt.year}_{dt.month:02d}.xlsx"


def save_expense(expense: Dict[str, Any]) -> Path:
    """
    Save or update a single expense into the month-year Excel file.
    - expense: dict containing at least 'id' and a date-like field ('date' preferred).
    - returns: Path to the Excel file written.
    Usage: call from your transactions create/update handlers:
        from .api.excel_backup import save_expense
        save_expense(expense_dict)
    """
    if not isinstance(expense, dict):
        raise ValueError("expense must be a dict")

    date_str = expense.get("date") or expense.get("created_at") or expense.get("updated_at")
    dt = _parse_date(date_str)
    path = _file_for_dt(dt)

    if path.exists():
        wb = load_workbook(filename=str(path))
        ws = wb.active
        existing_header = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
        if existing_header != HEADERS:
            # normalize header: clear sheet and recreate header
            ws.delete_rows(1, ws.max_row)
            ws.append(HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)

    exp_id = expense.get("id")
    row_idx = None
    if exp_id is not None:
        for row in ws.iter_rows(min_row=2):
            if row and row[0].value == exp_id:
                row_idx = row[0].row
                break

    if row_idx:
        for col_index, key in enumerate(HEADERS, start=1):
            ws.cell(row=row_idx, column=col_index, value=expense.get(key))
    else:
        row_values = [expense.get(k) for k in HEADERS]
        ws.append(row_values)

    wb.save(filename=str(path))
    LOG.info("Saved expense id=%s to %s", exp_id, path.name)
    return path
    wb.save(filename=str(path))
    LOG.info("Saved expense id=%s to %s", exp_id, path.name)
    return path
