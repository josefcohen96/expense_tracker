"""Backup and restoration utilities for the expense tracker.

This module provides functions to create Excel backups of the
entire database, list existing backups, enforce retention
policies and restore the database from a backup file. Excel
format is used for portability and ease of inspection by
non‑technical users. The `openpyxl` library (installed in this
environment) is used to read and write the XLSX files.
"""

from __future__ import annotations

import os
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Tuple

import openpyxl

from .db import get_connection


# Directory where backups will be stored
BASE_DIR = Path(__file__).resolve().parent
BACKUP_DIR = BASE_DIR.parent / "data" / "backups" / "excel"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

# Default number of backups to retain
DEFAULT_RETENTION = 30


def list_backups() -> List[Dict[str, Any]]:
    """Return a list of existing backup files sorted by creation time (descending)."""
    backups: List[Dict[str, Any]] = []
    for path in BACKUP_DIR.glob("*.xlsx"):
        stat = path.stat()
        backups.append({
            "file_name": path.name,
            "size": stat.st_size,
            "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        })
    backups.sort(key=lambda x: x["created_at"], reverse=True)
    return backups


def enforce_retention(retention: int = DEFAULT_RETENTION) -> None:
    """Delete old backup files, retaining only the most recent `retention` files."""
    backups = list_backups()
    if len(backups) <= retention:
        return
    # Determine files to delete (oldest first)
    to_delete = backups[retention:]
    for info in to_delete:
        try:
            (BACKUP_DIR / info["file_name"]).unlink()
        except FileNotFoundError:
            pass


def create_backup() -> Path:
    """Create a new Excel backup and return the file path.

    The backup contains one sheet per table in the database. A
    README sheet is added to describe the structure. A temporary
    file is written first and then atomically moved into place to
    avoid partial writes.
    """
    now = datetime.now()
    filename = f"budget_{now:%Y%m%d_%H%M%S}.xlsx"
    temp_path = BACKUP_DIR / f".{filename}.tmp"
    final_path = BACKUP_DIR / filename
    # Open connection and fetch all tables
    with get_connection() as conn:
        wb = openpyxl.Workbook()
        # Remove default sheet created by openpyxl
        wb.remove(wb.active)
        tables = ["users", "categories", "accounts", "recurrences", "transactions", "meta"]
        for table in tables:
            ws = wb.create_sheet(title=table)
            cur = conn.execute(f"SELECT * FROM {table}")
            cols = [desc[0] for desc in cur.description]
            ws.append(cols)
            for row in cur:
                ws.append(list(row))
            cur.close()
        # README sheet
        readme = wb.create_sheet(title="readme")
        readme.append([
            "This backup file contains one worksheet per database table."
        ])
        readme.append([
            "Restore by uploading this file back through the API."
        ])
        # Save to temporary file
        wb.save(temp_path)
    # Atomically move temp file to final location
    temp_path.rename(final_path)
    # Enforce retention
    enforce_retention()
    return final_path


def restore_from_backup(xlsx_path: Path) -> Tuple[int, int, int]:
    """Restore the database from an Excel backup.

    The function reads the given XLSX file, clears all existing
    records in the database tables and inserts the contents of
    each worksheet. It returns a tuple of counts of inserted
    rows for users, categories, accounts, recurrences,
    transactions and meta.

    Parameters
    ----------
    xlsx_path : pathlib.Path
        The path to the backup file.

    Returns
    -------
    Tuple[int, int, int, int, int, int]
        Counts of inserted rows in the order of tables.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=xlsx_path, data_only=True)
    # Ensure required sheets exist
    required = ["users", "categories", "accounts", "recurrences", "transactions", "meta"]
    for name in required:
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet '{name}' missing from backup")
    with get_connection() as conn:
        conn.execute("PRAGMA foreign_keys = OFF")
        # Clear tables
        for table in required:
            conn.execute(f"DELETE FROM {table}")
        inserted_counts = []
        # Insert data in order to satisfy foreign key dependencies
        for table in ["users", "categories", "accounts", "recurrences", "transactions", "meta"]:
            ws = wb[table]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                inserted_counts.append(0)
                continue
            headers = [str(h) for h in rows[0]]
            insert_count = 0
            for row in rows[1:]:
                # Build insert statement
                placeholders = ",".join(["?"] * len(headers))
                insert_sql = f"INSERT INTO {table} ({','.join(headers)}) VALUES ({placeholders})"
                conn.execute(insert_sql, tuple(row))
                insert_count += 1
            inserted_counts.append(insert_count)
        conn.execute("PRAGMA foreign_keys = ON")
        conn.commit()
        return tuple(inserted_counts)