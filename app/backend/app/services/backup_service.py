"""
Backup service for creating and managing database backups.
"""

from pathlib import Path
from datetime import datetime, date
from typing import Optional, List, Dict, Tuple
import sqlite3
import logging
import zipfile
import shutil
import os

from openpyxl import Workbook

LOG = logging.getLogger(__name__)

# Prevent re-entrant / recursive backups
_IN_PROGRESS = False

ROOT_DIR = Path(__file__).resolve().parents[2]  # .../expense_tracker/app
BACKUP_DIR = ROOT_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_ROOT = BACKUP_DIR / "excel"
EXCEL_ROOT.mkdir(parents=True, exist_ok=True)

# Headers for expenses table
EXPENSES_HEADERS = [
    "id", "date", "amount", "category", "user", "account", "notes", "tags", "recurrence_id"
]

# Headers for recurrences table (updated to next_charge_date)
RECURRENCES_HEADERS = [
    "id", "name", "amount", "category", "user", "frequency", "next_charge_date",
    "day_of_month", "weekday", "active"
]


def _last_n_months(n: int) -> List[Tuple[int, int]]:
    """Return list of (year, month) for current month and previous n-1 months."""
    today = date.today()
    result = []
    for i in range(n):
        total_month = today.year * 12 + today.month - 1 - i
        y = total_month // 12
        m = total_month % 12 + 1
        result.append((y, m))
    return list(reversed(result))  # older -> newer


def _find_db_file() -> Optional[Path]:
    """Find database file in known locations."""
    candidates = [
        os.getenv("BUDGET_DB_PATH"),  # preferred (set by tests/runner)
        os.getenv("COUPLEBUDGET_DB"),
        # common relative locations under project root
        "backend/app/data/budget.db",
        "app/backend/app/data/budget.db",
        "app/data/budget.db",
        "data/budget.db",
        "budget.db",
        "backend/app/data/couplebudget.sqlite3",
        "app/backend/app/data/couplebudget.sqlite3",
        "app/data/couplebudget.sqlite3",
        "data/couplebudget.sqlite3",
        "couplebudget.sqlite3",
        "data.db", "couplebudget.db", "db.sqlite3", "app.db", "database.db"
    ]
    for c in candidates:
        if not c:
            continue
        p = Path(c) if Path(c).is_absolute() else ROOT_DIR / c
        if p.exists():
            return p
    return None


def _open_conn_from_path_or_conn(db_conn: Optional[sqlite3.Connection]) -> sqlite3.Connection:
    """Open database connection from path or use provided connection."""
    if db_conn is not None:
        return db_conn
    db_path = _find_db_file()
    if not db_path:
        raise RuntimeError("No DB connection provided and DB file not found in known locations")
    return sqlite3.connect(str(db_path))


def create_backup_file(db_conn: Optional[sqlite3.Connection] = None) -> Path:
    """
    Create a dated folder (DD MM YYYY) under backups/ and write an Excel file for each of the
    last 6 months (including current) containing that month's transactions and recurrences.
    
    Args:
        db_conn: Optional database connection
    
    Returns:
        Path to the created folder
    """
    global _IN_PROGRESS
    if _IN_PROGRESS:
        raise RuntimeError("Backup already in progress (re-entrant call detected)")
    _IN_PROGRESS = True

    conn_provided = db_conn is not None
    conn = None
    try:
        conn = _open_conn_from_path_or_conn(db_conn)
        # ensure row factory for dict-like access
        try:
            conn.row_factory = sqlite3.Row
        except Exception:
            pass

        folder_name = datetime.utcnow().strftime("%d %m %Y")
        out_dir = BACKUP_DIR / folder_name
        out_dir.mkdir(parents=True, exist_ok=True)

        months = _last_n_months(6)
        for (year, month) in months:
            ym = f"{year}-{month:02d}"
            
            # Create workbook with two sheets
            wb = Workbook()
            
            # Remove default sheet and create our own
            wb.remove(wb.active)
            
            # Create expenses sheet
            expenses_ws = wb.create_sheet("הוצאות")
            expenses_ws.append(EXPENSES_HEADERS)
            
            # Get expenses for this month
            expenses_cur = conn.execute(
                """
                SELECT t.id, t.date, t.amount, c.name as category, u.name as user, 
                       a.name as account, t.notes, t.tags, t.recurrence_id
                FROM transactions t
                LEFT JOIN categories c ON t.category_id = c.id
                LEFT JOIN users u ON t.user_id = u.id
                LEFT JOIN accounts a ON t.account_id = a.id
                WHERE strftime('%Y-%m', t.date) = ?
                ORDER BY t.date ASC, t.id ASC
                """,
                (ym,),
            )
            expenses_rows = expenses_cur.fetchall()
            
            for r in expenses_rows:
                vals = [
                    r["id"],
                    r["date"],
                    r["amount"],
                    r["category"],
                    r["user"],
                    r["account"] or "",
                    r["notes"] or "",
                    r["tags"] or "",
                    r["recurrence_id"] or "",
                ]
                expenses_ws.append(vals)
            
            # Create recurrences sheet
            recurrences_ws = wb.create_sheet("הוצאות קבועות")
            recurrences_ws.append(RECURRENCES_HEADERS)
            
            # Get all active recurrences
            recurrences_cur = conn.execute(
                """
                SELECT r.id, r.name, r.amount, c.name as category, u.name as user,
                       r.frequency, r.next_charge_date, r.day_of_month, 
                       r.weekday, r.active
                FROM recurrences r
                LEFT JOIN categories c ON r.category_id = c.id
                LEFT JOIN users u ON r.user_id = u.id
                WHERE r.active = 1
                ORDER BY r.name ASC
                """,
            )
            recurrences_rows = recurrences_cur.fetchall()
            
            for r in recurrences_rows:
                vals = [
                    r["id"],
                    r["name"],
                    r["amount"],
                    r["category"],
                    r["user"],
                    r["frequency"],
                    r["next_charge_date"],
                    r["day_of_month"] or "",
                    r["weekday"] or "",
                    "כן" if r["active"] else "לא",
                ]
                recurrences_ws.append(vals)
            
            # Save the workbook
            file_name = f"monthly_backup_{year}_{month:02d}.xlsx"
            wb.save(filename=str(out_dir / file_name))

            LOG.info("Created excel backup folder %s", out_dir.name)
    finally:
        _IN_PROGRESS = False
        if conn is not None and not conn_provided:
            try:
                conn.close()
            except Exception:
                pass
    return out_dir


def create_monthly_backup(year: int, month: int, db_conn: Optional[sqlite3.Connection] = None) -> Path:
    """
    Create a single Excel file for a specific month with expenses and recurrences.
    
    Args:
        year: Year (e.g., 2024)
        month: Month (1-12)
        db_conn: Optional database connection
    
    Returns:
        Path to the created Excel file
    """
    global _IN_PROGRESS
    if _IN_PROGRESS:
        raise RuntimeError("Backup already in progress (re-entrant call detected)")
    _IN_PROGRESS = True

    conn_provided = db_conn is not None
    conn = None
    try:
        conn = _open_conn_from_path_or_conn(db_conn)
        # ensure row factory for dict-like access
        try:
            conn.row_factory = sqlite3.Row
        except Exception:
            pass

        ym = f"{year}-{month:02d}"
        
        # Create workbook with two sheets
        wb = Workbook()
        
        # Remove default sheet and create our own
        wb.remove(wb.active)
        
        # Create expenses sheet
        expenses_ws = wb.create_sheet("הוצאות")
        expenses_ws.append(EXPENSES_HEADERS)
        
        # Get expenses for this month
        expenses_cur = conn.execute(
            """
            SELECT t.id, t.date, t.amount, c.name as category, u.name as user, 
                   a.name as account, t.notes, t.tags, t.recurrence_id
            FROM transactions t
            LEFT JOIN categories c ON t.category_id = c.id
            LEFT JOIN users u ON t.user_id = u.id
            LEFT JOIN accounts a ON t.account_id = a.id
            WHERE strftime('%Y-%m', t.date) = ?
            ORDER BY t.date ASC, t.id ASC
            """,
            (ym,),
        )
        expenses_rows = expenses_cur.fetchall()
        
        for r in expenses_rows:
            vals = [
                r["id"],
                r["date"],
                r["amount"],
                r["category"],
                r["user"],
                r["account"] or "",
                r["notes"] or "",
                r["tags"] or "",
                r["recurrence_id"] or "",
            ]
            expenses_ws.append(vals)
        
        # Create recurrences sheet
        recurrences_ws = wb.create_sheet("הוצאות קבועות")
        recurrences_ws.append(RECURRENCES_HEADERS)
        
        # Get all active recurrences
        recurrences_cur = conn.execute(
            """
            SELECT r.id, r.name, r.amount, c.name as category, u.name as user,
                   r.frequency, r.next_charge_date, r.day_of_month, 
                   r.weekday, r.active
            FROM recurrences r
            LEFT JOIN categories c ON r.category_id = c.id
            LEFT JOIN users u ON r.user_id = u.id
            WHERE r.active = 1
            ORDER BY r.name ASC
            """,
        )
        recurrences_rows = recurrences_cur.fetchall()
        
        for r in recurrences_rows:
            vals = [
                r["id"],
                r["name"],
                r["amount"],
                r["category"],
                r["user"],
                r["frequency"],
                r["next_charge_date"],
                r["day_of_month"] or "",
                r["weekday"] or "",
                "כן" if r["active"] else "לא",
            ]
            recurrences_ws.append(vals)
        
        # Save the workbook
        file_name = f"monthly_backup_{year}_{month:02d}.xlsx"
        file_path = EXCEL_ROOT / file_name
        wb.save(filename=str(file_path))

        LOG.info("Created monthly backup file %s", file_name)
        return file_path
    finally:
        _IN_PROGRESS = False
        if conn is not None and not conn_provided:
            try:
                conn.close()
            except Exception:
                pass


def list_backup_files() -> List[Dict]:
    """
    List items in BACKUP_DIR. For files: file_name, created_at, size.
    For directories: treat the dir as a backup (name, mtime, aggregated size).
    """
    items = []
    for p in sorted(BACKUP_DIR.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            if p.is_file():
                stat = p.stat()
                items.append({
                    "file_name": p.name,
                    "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "size": stat.st_size,
                })
            elif p.is_dir():
                total = 0
                for f in p.rglob("*"):
                    if f.is_file():
                        try:
                            total += f.stat().st_size
                        except Exception:
                            continue
                stat = p.stat()
                items.append({
                    "file_name": p.name,
                    "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "size": total,
                })
        except Exception:
            LOG.exception("Failed to stat backup entry %s", p)
            continue
    return items


def restore_from_file(path: Path) -> Dict:
    """
    Restore backup from a path that can be either:
      - a zip file (will be extracted into EXCEL_ROOT)
      - a directory containing excel files (files will be copied into EXCEL_ROOT)
    Returns counts of files restored.
    """
    restored = {"files": 0}
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))

    if p.is_file() and zipfile.is_zipfile(str(p)):
        with zipfile.ZipFile(str(p), "r") as zf:
            tmp = BACKUP_DIR / f".restore_tmp_{datetime.utcnow().timestamp()}"
            if tmp.exists():
                shutil.rmtree(tmp)
            tmp.mkdir(parents=True, exist_ok=True)
            zf.extractall(tmp)
            for f in tmp.iterdir():
                dst = EXCEL_ROOT / f.name
                if dst.exists():
                    dst.unlink()
                shutil.move(str(f), str(dst))
                restored["files"] += 1
            shutil.rmtree(tmp)
    elif p.is_dir():
        for f in p.iterdir():
            if f.is_file() and f.suffix.lower() in (".xlsx", ".xlsm", ".xltx"):
                dst = EXCEL_ROOT / f.name
                if dst.exists():
                    dst.unlink()
                shutil.copy2(str(f), str(dst))
                restored["files"] += 1
    else:
        raise RuntimeError("Unsupported restore file type")

    LOG.info("Restored %d files from %s", restored["files"], p.name)
    return restored
