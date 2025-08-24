import sqlite3
from openpyxl import load_workbook
from pathlib import Path

def verify_real_backup():
    """Verify that the backup file contains the correct real data from the database."""
    
    # Database path
    db_path = "data/budget.db"
    backup_path = "backups/excel/monthly_backup_2025_08.xlsx"
    
    print("=== בדיקת גיבוי עם הנתונים האמיתיים ===")
    print(f"מסד נתונים: {db_path}")
    print(f"קובץ גיבוי: {backup_path}")
    print()
    
    # Connect to database
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    
    # Get data from database for August 2025
    print("=== נתונים ממסד הנתונים (אוגוסט 2025) ===")
    
    # Get transactions for August 2025
    transactions_query = """
        SELECT t.id, t.date, t.amount, c.name as category, u.name as user, 
               a.name as account, t.notes, t.tags, t.recurrence_id
        FROM transactions t
        LEFT JOIN categories c ON t.category_id = c.id
        LEFT JOIN users u ON t.user_id = u.id
        LEFT JOIN accounts a ON t.account_id = a.id
        WHERE strftime('%Y-%m', t.date) = '2025-08'
        ORDER BY t.date DESC, t.id DESC
    """
    
    transactions = conn.execute(transactions_query).fetchall()
    print(f"מספר עסקאות באוגוסט 2025: {len(transactions)}")
    
    # Get all active recurrences
    recurrences_query = """
        SELECT r.id, r.name, r.amount, c.name as category, u.name as user,
               r.frequency, r.start_date, r.end_date, r.day_of_month, 
               r.weekday, r.active
        FROM recurrences r
        LEFT JOIN categories c ON r.category_id = c.id
        LEFT JOIN users u ON r.user_id = u.id
        WHERE r.active = 1
        ORDER BY r.name ASC
    """
    
    recurrences = conn.execute(recurrences_query).fetchall()
    print(f"\nמספר הוצאות קבועות פעילות: {len(recurrences)}")
    
    conn.close()
    
    # Read backup file
    print("\n=== נתונים מקובץ הגיבוי ===")
    
    if Path(backup_path).exists():
        try:
            # Load workbook
            wb = load_workbook(backup_path)
            
            # Read expenses sheet
            if "הוצאות" in wb.sheetnames:
                expenses_ws = wb["הוצאות"]
                expenses_data = []
                
                # Skip header row
                for row in expenses_ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:  # Check if row has data
                        expenses_data.append({
                            'id': row[0],
                            'date': row[1],
                            'amount': row[2],
                            'category': row[3],
                            'user': row[4],
                            'account': row[5],
                            'notes': row[6],
                            'tags': row[7],
                            'recurrence_id': row[8]
                        })
                
                print(f"מספר עסקאות בקובץ הגיבוי: {len(expenses_data)}")
            else:
                print("❌ גיליון 'הוצאות' לא נמצא בקובץ הגיבוי")
            
            # Read recurrences sheet
            if "הוצאות קבועות" in wb.sheetnames:
                recurrences_ws = wb["הוצאות קבועות"]
                recurrences_data = []
                
                # Skip header row
                for row in recurrences_ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:  # Check if row has data
                        recurrences_data.append({
                            'id': row[0],
                            'name': row[1],
                            'amount': row[2],
                            'category': row[3],
                            'user': row[4],
                            'frequency': row[5],
                            'start_date': row[6],
                            'end_date': row[7],
                            'day_of_month': row[8],
                            'weekday': row[9],
                            'active': row[10]
                        })
                
                print(f"\nמספר הוצאות קבועות בקובץ הגיבוי: {len(recurrences_data)}")
            else:
                print("❌ גיליון 'הוצאות קבועות' לא נמצא בקובץ הגיבוי")
            
            # Verify data matches
            print("\n=== בדיקת התאמה ===")
            
            if len(transactions) == len(expenses_data):
                print("✅ מספר העסקאות תואם בין מסד הנתונים לקובץ הגיבוי")
            else:
                print(f"❌ מספר העסקאות לא תואם: מסד נתונים={len(transactions)}, גיבוי={len(expenses_data)}")
            
            if len(recurrences) == len(recurrences_data):
                print("✅ מספר ההוצאות הקבועות תואם בין מסד הנתונים לקובץ הגיבוי")
            else:
                print(f"❌ מספר ההוצאות הקבועות לא תואם: מסד נתונים={len(recurrences)}, גיבוי={len(recurrences_data)}")
            
        except Exception as e:
            print(f"❌ שגיאה בקריאת קובץ הגיבוי: {e}")
    else:
        print(f"❌ קובץ הגיבוי לא נמצא: {backup_path}")

if __name__ == "__main__":
    verify_real_backup()
