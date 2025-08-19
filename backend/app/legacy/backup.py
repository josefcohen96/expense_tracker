from __future__ import annotations

import os
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Tuple

import openpyxl

from .db import get_connection


BASE_DIR = Path(__file__).resolve().parent
BACKUP_DIR = BASE_DIR.parent.parent / "data" / "backups" / "excel"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

DEFAULT_RETENTION = 30


def _get_retention_days_from_env() -> int:
    try:
        value = int(os.getenv("BACKUP_RETENTION_DAYS", str(DEFAULT_RETENTION)))
        return value if value > 0 else DEFAULT_RETENTION
    except Exception:
        return DEFAULT_RETENTION


def has_backup_for_date(target_dt: datetime) -> bool:
    ymd = target_dt.strftime("%Y%m%d")
    for path in BACKUP_DIR.glob(f"budget_{ymd}_*.xlsx"):
        if path.exists():
            return True
    return False


def list_backups() -> List[Dict[str, Any]]:
    backups: List[Dict[str, Any]] = []
    for path in BACKUP_DIR.glob("*.xlsx"):
        stat = path.stat()
        backups.append({
            "file_name": path.name,
            "size": stat.st_size,
            "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        })
    backups.sort(key=lambda x: x["created_at"], reverse=True)
    return backups


def enforce_retention(retention: int = DEFAULT_RETENTION) -> None:
    backups = list_backups()
    if len(backups) <= retention:
        return
    to_delete = backups[retention:]
    for info in to_delete:
        try:
            (BACKUP_DIR / info["file_name"]).unlink()
        except FileNotFoundError:
            pass


def create_backup(only_if_no_backup_today: bool = False) -> Path:
    now = datetime.now()
    if only_if_no_backup_today and has_backup_for_date(now):
        existing = list_backups()
        for info in existing:
            if info["file_name"].startswith(now.strftime("budget_%Y%m%d_")):
                return BACKUP_DIR / info["file_name"]
    filename = f"budget_{now:%Y%m%d_%H%M%S}.xlsx"
    temp_path = BACKUP_DIR / f".{filename}.tmp"
    final_path = BACKUP_DIR / filename
    with get_connection() as conn:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        tables = ["users", "categories", "accounts", "recurrences", "transactions", "meta"]
        for table in tables:
            ws = wb.create_sheet(title=table)
            cur = conn.execute(f"SELECT * FROM {table}")
            cols = [desc[0] for desc in cur.description]
            ws.append(cols)
            for row in cur:
                ws.append(list(row))
            cur.close()
        readme = wb.create_sheet(title="readme")
        readme.append(["This backup file contains one worksheet per database table."])
        readme.append(["Restore by uploading this file back through the API."])
        wb.save(temp_path)
    temp_path.rename(final_path)
    enforce_retention(retention=_get_retention_days_from_env())
    return final_path


def restore_from_backup(xlsx_path: Path) -> Tuple[int, int, int]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=xlsx_path, data_only=True)
    required = ["users", "categories", "accounts", "recurrences", "transactions", "meta"]
    for name in required:
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet '{name}' missing from backup")
    with get_connection() as conn:
        conn.execute("PRAGMA foreign_keys = OFF")
        for table in required:
            conn.execute(f"DELETE FROM {table}")
        inserted_counts = []
        for table in ["users", "categories", "accounts", "recurrences", "transactions", "meta"]:
            ws = wb[table]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                inserted_counts.append(0)
                continue
            headers = [str(h) for h in rows[0]]
            insert_count = 0
            for row in rows[1:]:
                placeholders = ",".join(["?"] * len(headers))
                insert_sql = f"INSERT INTO {table} ({','.join(headers)}) VALUES ({placeholders})"
                conn.execute(insert_sql, tuple(row))
                insert_count += 1
            inserted_counts.append(insert_count)
        conn.execute("PRAGMA foreign_keys = ON")
        conn.commit()
        return tuple(inserted_counts)

