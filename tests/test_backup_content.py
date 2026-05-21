"""
Comprehensive backup content tests.
Verifies that Excel files produced by the backup service contain
exactly the right data — correct headers, correct rows, correct values —
and that the file is a valid, openable Excel workbook.
"""

import io
import sqlite3
import shutil
import zipfile
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.backend.app.services.backup_service import (
    create_backup_file,
    create_monthly_backup,
    BACKUP_DIR,
    EXCEL_ROOT,
    EXPENSES_HEADERS,
    RECURRENCES_HEADERS,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _setup_isolated_db(tmp_path: Path) -> sqlite3.Connection:
    """
    Create a fresh in-memory-style SQLite DB at tmp_path with all required
    tables and known seed data. Returns an open connection.
    """
    db_path = tmp_path / "test_backup.db"
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")

    conn.executescript("""
        CREATE TABLE categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            is_saving BOOLEAN DEFAULT 0
        );
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );
        CREATE TABLE accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );
        CREATE TABLE recurrences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            amount REAL NOT NULL,
            category_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            frequency TEXT NOT NULL,
            day_of_month INTEGER,
            weekday INTEGER,
            next_charge_date TEXT NOT NULL,
            active BOOLEAN DEFAULT 1
        );
        CREATE TABLE transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            amount REAL NOT NULL,
            category_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            account_id INTEGER,
            notes TEXT,
            tags TEXT,
            recurrence_id INTEGER,
            period_key TEXT
        );
    """)

    # Seed lookup tables
    conn.execute("INSERT INTO categories (id, name) VALUES (1, 'אוכל')")
    conn.execute("INSERT INTO categories (id, name) VALUES (2, 'בריאות')")
    conn.execute("INSERT INTO users    (id, name) VALUES (1, 'Yosef')")
    conn.execute("INSERT INTO users    (id, name) VALUES (2, 'Karina')")
    conn.execute("INSERT INTO accounts (id, name) VALUES (1, 'כרטיס אשראי')")
    conn.execute("INSERT INTO accounts (id, name) VALUES (2, 'מזומן')")

    # One active recurrence
    conn.execute("""
        INSERT INTO recurrences (id, name, amount, category_id, user_id, frequency,
                                 day_of_month, next_charge_date, active)
        VALUES (1, 'Netflix', 50.0, 1, 1, 'monthly', 5, '2026-06-05', 1)
    """)
    # One inactive recurrence (should NOT appear in backup)
    conn.execute("""
        INSERT INTO recurrences (id, name, amount, category_id, user_id, frequency,
                                 day_of_month, next_charge_date, active)
        VALUES (2, 'Spotify', 20.0, 1, 2, 'monthly', 10, '2026-06-10', 0)
    """)

    conn.commit()
    return conn


def _insert_transaction(conn, *, date_str, amount, cat_id=1, user_id=1,
                         account_id=None, notes=None, tags=None, rec_id=None, period_key=None):
    conn.execute(
        """INSERT INTO transactions (date, amount, category_id, user_id, account_id,
                                     notes, tags, recurrence_id, period_key)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (date_str, amount, cat_id, user_id, account_id, notes, tags, rec_id, period_key),
    )
    conn.commit()
    return conn.execute("SELECT last_insert_rowid()").fetchone()[0]


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestExcelFileValidity:
    """The produced file must be a valid, openable Excel workbook."""

    def test_monthly_backup_opens_as_valid_xlsx(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        _insert_transaction(conn, date_str=today.isoformat(), amount=100.0)

        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        assert out.exists(), "Backup file was not created"
        assert out.suffix == ".xlsx", "File must have .xlsx extension"

        # openpyxl raises if the file is corrupt / not a valid OOXML package
        wb = load_workbook(filename=str(out))
        assert wb is not None

    def test_full_backup_zip_contains_six_xlsx_files(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        _insert_transaction(conn, date_str=today.isoformat(), amount=200.0)

        zip_path = create_backup_file(db_conn=conn)
        conn.close()

        assert zip_path.suffix == ".zip", f"Expected a .zip file, got: {zip_path}"
        assert zip_path.exists(), "ZIP file was not created"

        with zipfile.ZipFile(str(zip_path), "r") as zf:
            names = zf.namelist()
            xlsx_names = [n for n in names if n.endswith(".xlsx")]
            assert len(xlsx_names) == 6, (
                f"Expected 6 xlsx files inside zip, found {len(xlsx_names)}: {xlsx_names}"
            )
            # Verify each member is a valid workbook
            for name in xlsx_names:
                data = zf.read(name)
                import io
                wb = load_workbook(filename=io.BytesIO(data))
                assert wb is not None, f"{name} inside zip is not a valid workbook"


class TestSheetNamesAndHeaders:
    """Correct sheet names and column headers."""

    def _load_current_month_backup(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        _insert_transaction(conn, date_str=today.isoformat(), amount=1.0)
        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()
        return load_workbook(filename=str(out))

    def test_sheet_names_are_correct(self, tmp_path):
        wb = self._load_current_month_backup(tmp_path)
        assert "הוצאות" in wb.sheetnames, "Missing sheet 'הוצאות'"
        assert "הוצאות קבועות" in wb.sheetnames, "Missing sheet 'הוצאות קבועות'"

    def test_expenses_headers(self, tmp_path):
        wb = self._load_current_month_backup(tmp_path)
        ws = wb["הוצאות"]
        actual = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert actual == EXPENSES_HEADERS, f"Headers mismatch: {actual}"

    def test_recurrences_headers(self, tmp_path):
        wb = self._load_current_month_backup(tmp_path)
        ws = wb["הוצאות קבועות"]
        actual = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert actual == RECURRENCES_HEADERS, f"Headers mismatch: {actual}"


def _is_empty(val) -> bool:
    """openpyxl returns None for blank cells regardless of whether '' or None was written."""
    return val is None or val == ""


class TestExpensesContent:
    """Verifies that expense rows match what's in the DB."""

    def test_all_columns_present_and_correct(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        tx_id = _insert_transaction(
            conn,
            date_str=today.isoformat(),
            amount=123.45,
            cat_id=1,
            user_id=2,           # Karina
            account_id=1,        # כרטיס אשראי
            notes="הוצאה לבדיקה",
            tags="חשוב,בדיקה",
        )

        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        wb = load_workbook(filename=str(out))
        ws = wb["הוצאות"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 1, f"Expected 1 expense row, got {len(rows)}"
        row = rows[0]

        assert row[0] == tx_id,          f"id: {row[0]} != {tx_id}"
        assert row[1] == today.isoformat(), f"date: {row[1]}"
        assert abs(row[2] - 123.45) < 0.001, f"amount: {row[2]}"
        assert row[3] == "אוכל",         f"category: {row[3]}"
        assert row[4] == "Karina",        f"user: {row[4]}"
        assert row[5] == "כרטיס אשראי",  f"account: {row[5]}"
        assert row[6] == "הוצאה לבדיקה", f"notes: {row[6]}"
        assert row[7] == "חשוב,בדיקה",   f"tags: {row[7]}"
        # recurrence_id is NULL → openpyxl reads blank cells as None
        assert _is_empty(row[8]), f"recurrence_id should be blank, got: {row[8]!r}"

    def test_recurrence_id_populated_when_linked(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        tx_id = _insert_transaction(
            conn,
            date_str=today.isoformat(),
            amount=50.0,
            rec_id=1,
            period_key="2026-05",
        )

        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        wb = load_workbook(filename=str(out))
        ws = wb["הוצאות"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][8] == 1, f"recurrence_id should be 1, got {rows[0][8]}"

    def test_null_optional_fields_are_blank_in_excel(self, tmp_path):
        """NULL fields in DB must appear as blank cells in Excel (not crash or show 'None')."""
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        _insert_transaction(
            conn,
            date_str=today.isoformat(),
            amount=10.0,
            account_id=None,
            notes=None,
            tags=None,
            rec_id=None,
        )

        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        wb = load_workbook(filename=str(out))
        rows = list(wb["הוצאות"].iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        row = rows[0]
        # openpyxl returns None for blank cells; any falsy value is acceptable here
        assert _is_empty(row[5]), f"account should be blank, got {row[5]!r}"
        assert _is_empty(row[6]), f"notes should be blank, got {row[6]!r}"
        assert _is_empty(row[7]), f"tags should be blank, got {row[7]!r}"
        assert _is_empty(row[8]), f"recurrence_id should be blank, got {row[8]!r}"

    def test_multiple_transactions_all_appear_ordered_by_date(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        ym = f"{today.year}-{today.month:02d}"

        dates_and_amounts = [
            (f"{ym}-03", 30.0),
            (f"{ym}-01", 10.0),
            (f"{ym}-05", 50.0),
            (f"{ym}-02", 20.0),
        ]
        for d, a in dates_and_amounts:
            _insert_transaction(conn, date_str=d, amount=a)

        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        wb = load_workbook(filename=str(out))
        rows = list(wb["הוצאות"].iter_rows(min_row=2, values_only=True))

        assert len(rows) == 4, f"Expected 4 rows, got {len(rows)}"
        # Must be sorted ascending by date
        extracted_dates = [r[1] for r in rows]
        assert extracted_dates == sorted(extracted_dates), (
            f"Rows not sorted by date: {extracted_dates}"
        )
        # Each amount must appear exactly once
        amounts = sorted(r[2] for r in rows)
        assert amounts == [10.0, 20.0, 30.0, 50.0]

    def test_only_current_month_transactions_included(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        ym = f"{today.year}-{today.month:02d}"

        _insert_transaction(conn, date_str=f"{ym}-10",    amount=100.0)   # this month ✓
        _insert_transaction(conn, date_str="2020-01-01",  amount=999.0)   # past ✗
        _insert_transaction(conn, date_str="2030-12-31",  amount=888.0)   # future ✗

        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        rows = list(load_workbook(filename=str(out))["הוצאות"].iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1, f"Expected 1 row, got {len(rows)}: {rows}"
        assert abs(rows[0][2] - 100.0) < 0.001

    def test_large_amount_precision(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        _insert_transaction(conn, date_str=today.isoformat(), amount=9999.99)

        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        rows = list(load_workbook(filename=str(out))["הוצאות"].iter_rows(min_row=2, values_only=True))
        assert abs(rows[0][2] - 9999.99) < 0.001, f"Precision lost: {rows[0][2]}"

    def test_hebrew_text_preserved_correctly(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        _insert_transaction(
            conn,
            date_str=today.isoformat(),
            amount=1.0,
            notes="עברית: שלום עולם! @#$%",
            tags="תג-אחד,תג-שניים",
        )

        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        rows = list(load_workbook(filename=str(out))["הוצאות"].iter_rows(min_row=2, values_only=True))
        assert rows[0][6] == "עברית: שלום עולם! @#$%", f"Hebrew notes corrupted: {rows[0][6]}"
        assert rows[0][7] == "תג-אחד,תג-שניים", f"Hebrew tags corrupted: {rows[0][7]}"

    def test_empty_month_has_only_headers(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        # Use a month in the past with no transactions
        out = create_monthly_backup(2021, 1, db_conn=conn)
        conn.close()

        ws = load_workbook(filename=str(out))["הוצאות"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert rows == [], f"Expected no data rows, found: {rows}"


class TestRecurrencesContent:
    """Verifies the recurrences sheet is correct."""

    def test_only_active_recurrences_appear(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        ws = load_workbook(filename=str(out))["הוצאות קבועות"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # Spotify (inactive) must NOT appear; Netflix (active) must appear
        names = [r[1] for r in rows]
        assert "Netflix" in names, f"Active recurrence missing: {names}"
        assert "Spotify" not in names, f"Inactive recurrence must not appear: {names}"

    def test_recurrence_row_values(self, tmp_path):
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        out = create_monthly_backup(today.year, today.month, db_conn=conn)
        conn.close()

        ws = load_workbook(filename=str(out))["הוצאות קבועות"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1

        row = rows[0]
        assert row[0] == 1,           f"id: {row[0]}"
        assert row[1] == "Netflix",   f"name: {row[1]}"
        assert abs(row[2] - 50.0) < 0.001, f"amount: {row[2]}"
        assert row[3] == "אוכל",      f"category: {row[3]}"
        assert row[4] == "Yosef",     f"user: {row[4]}"
        assert row[5] == "monthly",   f"frequency: {row[5]}"
        assert row[6] == "2026-06-05", f"next_charge_date: {row[6]}"
        assert row[7] == 5,           f"day_of_month: {row[7]}"
        assert _is_empty(row[8]),     f"weekday should be blank (None/NULL), got {row[8]!r}"
        assert row[9] == "כן",        f"active flag: {row[9]}"

    def test_recurrences_same_in_every_month_of_full_backup(self, tmp_path):
        """Full backup ZIP (6 months): recurrences sheet must be identical in each file."""
        import io
        conn = _setup_isolated_db(tmp_path)
        zip_path = create_backup_file(db_conn=conn)
        conn.close()

        all_rec_rows = []
        with zipfile.ZipFile(str(zip_path), "r") as zf:
            for name in sorted(zf.namelist()):
                wb = load_workbook(filename=io.BytesIO(zf.read(name)))
                rows = list(wb["הוצאות קבועות"].iter_rows(min_row=2, values_only=True))
                all_rec_rows.append((name, rows))

        assert all_rec_rows, "ZIP must contain at least one xlsx"
        first_rows = all_rec_rows[0][1]
        for fname, rows in all_rec_rows[1:]:
            assert rows == first_rows, (
                f"{fname} recurrences differ from {all_rec_rows[0][0]}"
            )


class TestFullBackupMonthCoverage:
    """Verifies the 6-month full backup ZIP covers the right months."""

    def _zip_names(self, tmp_path) -> set:
        conn = _setup_isolated_db(tmp_path)
        zip_path = create_backup_file(db_conn=conn)
        conn.close()
        with zipfile.ZipFile(str(zip_path), "r") as zf:
            return set(zf.namelist())

    def test_file_names_cover_last_six_months(self, tmp_path):
        today = date.today()
        expected = set()
        for i in range(6):
            total = today.year * 12 + today.month - 1 - i
            y = total // 12
            m = total % 12 + 1
            expected.add(f"monthly_backup_{y}_{m:02d}.xlsx")

        actual = self._zip_names(tmp_path)
        assert actual == expected, (
            f"ZIP contents mismatch.\nExpected: {expected}\nGot: {actual}"
        )

    def test_transactions_appear_only_in_correct_month_file(self, tmp_path):
        import io
        conn = _setup_isolated_db(tmp_path)
        today = date.today()
        ym = f"{today.year}-{today.month:02d}"
        _insert_transaction(conn, date_str=f"{ym}-10", amount=777.0)

        zip_path = create_backup_file(db_conn=conn)
        conn.close()

        current_name = f"monthly_backup_{today.year}_{today.month:02d}.xlsx"
        with zipfile.ZipFile(str(zip_path), "r") as zf:
            assert current_name in zf.namelist(), f"{current_name} not found in zip"

            # Current month must have the transaction
            wb_cur = load_workbook(filename=io.BytesIO(zf.read(current_name)))
            cur_rows = list(wb_cur["הוצאות"].iter_rows(min_row=2, values_only=True))
            assert len(cur_rows) == 1 and abs(cur_rows[0][2] - 777.0) < 0.001

            # Other months must be empty
            for name in zf.namelist():
                if name == current_name:
                    continue
                wb = load_workbook(filename=io.BytesIO(zf.read(name)))
                rows = list(wb["הוצאות"].iter_rows(min_row=2, values_only=True))
                assert rows == [], f"{name} unexpectedly contains rows: {rows}"


class TestReentryGuard:
    """Re-entrant call must raise RuntimeError."""

    def test_monthly_backup_reentrancy_blocked(self, tmp_path):
        import app.backend.app.services.backup_service as svc
        conn = _setup_isolated_db(tmp_path)
        today = date.today()

        original = svc._IN_PROGRESS
        svc._IN_PROGRESS = True
        try:
            with pytest.raises(RuntimeError, match="already in progress"):
                create_monthly_backup(today.year, today.month, db_conn=conn)
        finally:
            svc._IN_PROGRESS = original
            conn.close()
